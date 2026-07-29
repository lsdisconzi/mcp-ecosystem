import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# Parse the data from the text
data = """
12/01/2026 003 0000059008 RECTO DOCTO:0000000488-02/03 - INOVA NUTRICAO ANIMAL LTDA - 49.946.126/0001-42 - DOCTO REF PEDIDO DE VENDA NRO: 256
21/01/2026 003 0000059008 RECTO DOCTO:0000000488-03/03 - INOVA NUTRICAO ANIMAL LTDA - 49.946.126/0001-42 - DOCTO REF PEDIDO DE VENDA NRO: 256
09/01/2026 003 0000059008 RECTO DOCTO:0000000490-01/01 - RESOLVE FOOD TRADING LLC - -
20/02/2026 003 0000059008 RECTO DOCTO:0000000513-01/01 - RESOLVE FOOD TRADING LLC - -
09/01/2026 003 0000059008 RECTO DOCTO:0000000490-01/01 - RESOLVE FOOD TRADING LLC - -
08/01/2026 003 0000059008 RECTO DOCTO:0000000508-01/01 - PORTO RECICLAGEM ANIMAL LTDA - 03.124.463/0001-53 - DOCTO REF PEDIDO DE VENDA NRO: 269
18/03/2026 003 0000059008 RECTO DOCTO:0000000509-01/01 - RESOLVE FOOD TRADING LLC - - DOCTO REF PEDIDO DE VENDA NRO: 257
02/02/2026 002 0000067677 RECTOS
15/01/2026 003 0000059008 RECTO DOCTO:0000000518-01/01 - PORTO RECICLAGEM ANIMAL LTDA - 03.124.463/0001-53 - DOCTO REF PEDIDO DE VENDA NRO: 274
09/01/2026 003 0000059008 RECTO DOCTO:0000000490-01/01 - RESOLVE FOOD TRADING LLC - -
16/03/2026 003 0000059008 RECTO DOCTO:0000000524-02/03 - INOVA NUTRICAO ANIMAL LTDA - 49.946.126/0001-42 - DOCTO REF PEDIDO DE VENDA NRO: 259
02/03/2026 003 0000059008 RECTO DOCTO:0000000524-01/03 - INOVA NUTRICAO ANIMAL LTDA - 49.946.126/0001-42 - DOCTO REF PEDIDO DE VENDA NRO: 259
24/03/2026 003 0000059008 RECTO DOCTO:0000000524-02/03 - INOVA NUTRICAO ANIMAL LTDA - 49.946.126/0001-42 - DOCTO REF PEDIDO DE VENDA NRO: 259
06/02/2026 003 0000059008 RECTO DOCTO:0000000529-01/01 - TO CENTRO OESTE OLEO QUIMICA LTDA - 08.251.405/0005-38 - DOCTO REF PEDIDO DE VENDA NRO: 278
06/02/2026 003 0000059008 RECTO DOCTO:0000000529-01/01 - TO CENTRO OESTE OLEO QUIMICA LTDA - 08.251.405/0005-38 - DOCTO REF PEDIDO DE VENDA NRO: 278
10/03/2026 003 0000059008 RECTO DOCTO:0000000532-01/01 - RESOLVE FOOD TRADING LLC - - DOCTO REF PEDIDO DE VENDA NRO: 271
06/02/2026 003 0000059008 RECTO DOCTO:0000000529-01/01 - TO CENTRO OESTE OLEO QUIMICA LTDA - 08.251.405/0005-38 - DOCTO REF PEDIDO DE VENDA NRO: 278
02/02/2026 002 0000067677 RECTOS
10/03/2026 003 0000059008 RECTO DOCTO:0000000543-02/03 - PET MANIA COMERCIO INTERNACIONAL LTDA - 12.494.512/0001-30 - DOCTO REF PEDIDO DE VENDA NRO: 277
19/02/2026 003 0000059008 RECTO DOCTO:0000000543-01/03 - PET MANIA COMERCIO INTERNACIONAL LTDA - 12.494.512/0001-30 - DOCTO REF PEDIDO DE VENDA NRO: 277
02/02/2026 002 0000067677 RECTOS
11/02/2026 002 0000067677 RECTO DOCTO:0000000546-01/01 - RESOLVE FOOD TRADING LLC - -
03/03/2026 002 0000067677 RECTOS
19/02/2026 003 0000059008 RECTO DOCTO:0000000547-01/01 - PORTO RECICLAGEM ANIMAL LTDA - 03.124.463/0001-53 - DOCTO REF PEDIDO DE VENDA NRO: 285
16/03/2026 003 0000059008 RECTOS
03/03/2026 002 0000067677 RECTOS
03/03/2026 002 0000067677 RECTOS
16/03/2026 003 0000059008 RECTOS
"""

data_prazo = """
09/01/2026 003 0000059008 RECTO DOCTO:0000000490-01/01 - RESOLVE FOOD TRADING LLC - -
07/01/2026 003 0000059008 RECTO DOCTO:0000000490-01/01 - RESOLVE FOOD TRADING LLC - -
20/01/2026 002 0000067677 RECTO DOCTO:0000000513-01/01 - RESOLVE FOOD TRADING LLC - -
14/01/2026 003 0000059008 RECTO DOCTO:0000000512-01/01 - FREBO PET ALIMENTOS PARA ANIMAIS LTDA - 41.375.216/0001-19 - DOCTO REF PEDIDO DE VENDA NRO: 273
12/02/2026 003 0000059008 RECTO DOCTO:0000000512-01/01 - FREBO PET ALIMENTOS PARA ANIMAIS LTDA - 41.375.216/0001-19 - DOCTO REF PEDIDO DE VENDA NRO: 273
20/02/2026 003 0000059008 RECTO DOCTO:0000000513-01/01 - RESOLVE FOOD TRADING LLC - -
22/01/2026 003 0000059008 RECTO DOCTO:0000000536-01/01 - QUANTO ALIMENTOS INDUSTRIA E COMERCIO LTDA - 05.895.634/0001-73 - DOCTO REF PEDIDO DE VENDA NRO: 276
23/01/2026 003 0000059008 RECTO DOCTO:0000000537-01/01 - IRMAOS VERONEZE LTDA - EPP - 07.544.140/0001-70 - DOCTO REF PEDIDO DE VENDA NRO: 281
10/02/2026 003 0000059008 RECTO DOCTO:0000000548-01/01 - RESOLVE FOOD TRADING LLC - - DOCTO REF PEDIDO DE VENDA NRO: 283
20/02/2026 003 0000059008 RECTO DOCTO:0000000513-01/01 - RESOLVE FOOD TRADING LLC - -
20/02/2026 003 0000059008 RECTO DOCTO:0000000513-01/01 - RESOLVE FOOD TRADING LLC - -
03/03/2026 002 0000067677 RECTOS
"""

def parse_transactions(text, class_name):
    transactions = []
    lines = text.strip().split('\n')
    
    for line in lines:
        if line.strip() and not line.startswith('Emitido') and not line.startswith('Total'):
            parts = line.split()
            if len(parts) >= 6:
                # Extract data
                data_docto = parts[0] if parts[0] else ''
                banco = parts[1] if len(parts) > 1 else ''
                conta = parts[2] if len(parts) > 2 else ''
                documento = parts[3] if len(parts) > 3 else ''
                # Join remaining as histórico
                historico = ' '.join(parts[4:]) if len(parts) > 4 else ''
                
                transactions.append({
                    'Classe': class_name,
                    'Data Doc.': data_docto,
                    'Banco': banco,
                    'Conta': conta,
                    'Documento': documento,
                    'Histórico': historico,
                    'Valor a débito': ''
                })
    return transactions

# Parse both classes
transactions_vista = parse_transactions(data, "9 - VENDAS A VISTA")
transactions_prazo = parse_transactions(data_prazo, "6 - VENDAS A PRAZO")

# Combine all transactions
all_transactions = transactions_vista + transactions_prazo

# Create DataFrame
df = pd.DataFrame(all_transactions)

# Create Excel file with styling
wb = Workbook()
ws = wb.active
ws.title = "Lançamentos Financeiros"

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add header row
headers = ['Classe', 'Data Doc.', 'Banco', 'Conta', 'Documento', 'Histórico', 'Valor a débito']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Add data rows
for row_num, transaction in enumerate(all_transactions, 2):
    ws.cell(row=row_num, column=1, value=transaction['Classe']).border = border
    ws.cell(row=row_num, column=2, value=transaction['Data Doc.']).border = border
    ws.cell(row=row_num, column=3, value=transaction['Banco']).border = border
    ws.cell(row=row_num, column=4, value=transaction['Conta']).border = border
    ws.cell(row=row_num, column=5, value=transaction['Documento']).border = border
    ws.cell(row=row_num, column=6, value=transaction['Histórico']).border = border
    ws.cell(row=row_num, column=7, value=transaction['Valor a débito']).border = border

# Add summary section
summary_start_row = len(all_transactions) + 3
ws.cell(row=summary_start_row, column=1, value="RESUMO POR CLASSE").font = Font(bold=True, size=12)
summary_start_row += 1

# Count by class
vista_count = len(transactions_vista)
prazo_count = len(transactions_prazo)

ws.cell(row=summary_start_row, column=1, value="9 - VENDAS A VISTA:")
ws.cell(row=summary_start_row, column=2, value=f"{vista_count} lançamentos")
summary_start_row += 1
ws.cell(row=summary_start_row, column=1, value="6 - VENDAS A PRAZO:")
ws.cell(row=summary_start_row, column=2, value=f"{prazo_count} lançamentos")
summary_start_row += 1
ws.cell(row=summary_start_row, column=1, value="TOTAL:")
ws.cell(row=summary_start_row, column=2, value=f"{len(all_transactions)} lançamentos")

# Add company info header
ws.cell(row=1, column=8, value="Empresa: 0001 - IBSCO - INDUSTRIA BRASILEIRA STIVAL COMP.")
ws.cell(row=2, column=8, value="Filial: 0001 - MATRIZ")
ws.cell(row=3, column=8, value="Período: 01/01/2026 até 31/03/2026")
ws.cell(row=4, column=8, value="Emitido em: 31/03/2026")

# Adjust column widths
column_widths = [20, 12, 8, 15, 15, 50, 15]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = width

# Auto-filter
ws.auto_filter.ref = f"A1:G{len(all_transactions) + 1}"

# Freeze header row
ws.freeze_panes = "A2"

# Save the file
output_file = "lancamentos_financeiros.xlsx"
wb.save(output_file)

print(f"✅ Arquivo Excel criado com sucesso: {output_file}")
print(f"📊 Total de lançamentos: {len(all_transactions)}")
print(f"   - Vendas à vista: {vista_count}")
print(f"   - Vendas a prazo: {prazo_count}")